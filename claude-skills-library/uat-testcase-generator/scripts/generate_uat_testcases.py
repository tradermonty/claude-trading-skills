#!/usr/bin/env python3
"""
UAT Test Case Generator Script

This script generates UAT test case Excel files with:
1. Summary sheet with category breakdown and progress tracking
2. Full test cases sheet with all test case details

Usage:
    python generate_uat_testcases.py \\
        --project "Redac Salesforce CRM" \\
        --output "UAT受入試験テストケース一覧.xlsx" \\
        --testcases testcases.json

Input Format (testcases.json):
    [
        {
            "id": "DATA-001",
            "priority": "High",
            "category": "データ移行検証",
            "sub_category": "移行データ件数照合",
            "scenario": "Account件数照合",
            "precondition": "既存CRMからデータ移行済み",
            "steps": "1. 既存CRMのAccount総件数を確認\\n2. Salesforce UAT環境のAccount総件数を確認\\n3. 件数を比較",
            "expected": "移行元と移行先のAccount件数が一致すること",
            "criteria": "件数差異0件"
        }
    ]
"""

import json
import argparse
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def parse_arguments():
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(description='Generate UAT test case Excel file')
    parser.add_argument('--project', required=True, help='Project name')
    parser.add_argument('--output', required=True, help='Output Excel file path')
    parser.add_argument('--testcases', required=True, help='Test cases JSON file path')
    return parser.parse_args()


def load_testcases(json_path):
    """Load test cases from JSON file."""
    with open(json_path, 'r', encoding='utf-8') as f:
        return json.load(f)


def create_summary_sheet(testcases):
    """Create summary sheet with category breakdown and progress."""
    # Count by category and priority
    category_counts = {}
    for tc in testcases:
        category = tc['category']
        priority = tc['priority']

        if category not in category_counts:
            category_counts[category] = {'total': 0, 'High': 0, 'Medium': 0, 'Low': 0}

        category_counts[category]['total'] += 1
        category_counts[category][priority] += 1

    # Create summary data
    summary_data = []

    # Title row
    summary_data.append(['UAT受入試験 テストケース一覧', None, None, None, None, None, None, None])
    summary_data.append([None, None, None, None, None, None, None, None])

    # Category counts section
    summary_data.append(['カテゴリ別テスト項目数', None, None, None, None, None, None, None])
    summary_data.append(['カテゴリ', '項目数', '優先度High', '優先度Medium', '優先度Low', None, None, None])

    for category, counts in category_counts.items():
        summary_data.append([
            category,
            counts['total'],
            counts['High'],
            counts['Medium'],
            counts['Low'],
            None, None, None
        ])

    # Total row
    total_count = sum(cat['total'] for cat in category_counts.values())
    total_high = sum(cat['High'] for cat in category_counts.values())
    total_medium = sum(cat['Medium'] for cat in category_counts.values())
    total_low = sum(cat['Low'] for cat in category_counts.values())
    summary_data.append(['合計', total_count, total_high, total_medium, total_low, None, None, None])

    summary_data.append([None, None, None, None, None, None, None, None])

    # Progress section
    summary_data.append(['進捗状況', None, None, None, None, None, None, None])
    summary_data.append(['カテゴリ', '項目数', '実施済み', '未実施', 'Pass', 'Fail', 'Pending', '進捗率'])

    for category, counts in category_counts.items():
        summary_data.append([
            category,
            counts['total'],
            0,  # 実施済み (initially 0)
            counts['total'],  # 未実施 (initially all)
            0,  # Pass (initially 0)
            0,  # Fail (initially 0)
            0,  # Pending (initially 0)
            0   # 進捗率 (initially 0)
        ])

    # Total row for progress
    summary_data.append([
        '合計',
        total_count,
        0,
        total_count,
        0,
        0,
        0,
        None
    ])

    return pd.DataFrame(summary_data)


def create_testcases_sheet(testcases):
    """Create full test cases sheet."""
    # Columns for test cases
    columns = [
        'テストケースID', '優先度', 'カテゴリ', 'サブカテゴリ', 'テストシナリオ',
        '事前条件', 'テスト手順', '期待結果', '実際の結果', '合格基準',
        '合否判定', '実施日', '実施者', '備考', '不具合ID'
    ]

    # Create rows
    rows = []
    for tc in testcases:
        rows.append([
            tc['id'],
            tc['priority'],
            tc['category'],
            tc['sub_category'],
            tc['scenario'],
            tc['precondition'],
            tc['steps'],
            tc['expected'],
            '',  # 実際の結果 (empty)
            tc['criteria'],
            '',  # 合否判定 (empty)
            '',  # 実施日 (empty)
            '',  # 実施者 (empty)
            '',  # 備考 (empty)
            ''   # 不具合ID (empty)
        ])

    return pd.DataFrame(rows, columns=columns)


def apply_formatting(file_path):
    """Apply Excel formatting to the generated file."""
    wb = load_workbook(file_path)

    # Format summary sheet
    ws_summary = wb['サマリー']

    # Header formatting
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)

    # Apply header formatting to row 4 and row 14
    for row in [4, 14]:
        for col in range(1, 9):
            cell = ws_summary.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Title formatting
    ws_summary['A1'].font = Font(bold=True, size=14)

    # Section title formatting
    section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    section_font = Font(bold=True, size=11)
    for cell in [ws_summary['A3'], ws_summary['A13']]:
        cell.fill = section_fill
        cell.font = section_font

    # Format test cases sheet
    ws_testcases = wb['全テストケース']

    # Header row formatting
    for col in range(1, 16):
        cell = ws_testcases.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Set column widths
    ws_testcases.column_dimensions['A'].width = 15  # テストケースID
    ws_testcases.column_dimensions['B'].width = 10  # 優先度
    ws_testcases.column_dimensions['C'].width = 20  # カテゴリ
    ws_testcases.column_dimensions['D'].width = 20  # サブカテゴリ
    ws_testcases.column_dimensions['E'].width = 30  # テストシナリオ
    ws_testcases.column_dimensions['F'].width = 30  # 事前条件
    ws_testcases.column_dimensions['G'].width = 40  # テスト手順
    ws_testcases.column_dimensions['H'].width = 30  # 期待結果
    ws_testcases.column_dimensions['I'].width = 30  # 実際の結果
    ws_testcases.column_dimensions['J'].width = 20  # 合格基準
    ws_testcases.column_dimensions['K'].width = 10  # 合否判定
    ws_testcases.column_dimensions['L'].width = 12  # 実施日
    ws_testcases.column_dimensions['M'].width = 12  # 実施者
    ws_testcases.column_dimensions['N'].width = 30  # 備考
    ws_testcases.column_dimensions['O'].width = 15  # 不具合ID

    # Apply borders and alignment to all cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws_testcases.iter_rows(min_row=1, max_row=ws_testcases.max_row,
                                      min_col=1, max_col=15):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Freeze header row
    ws_testcases.freeze_panes = 'A2'

    wb.save(file_path)


def main():
    """Main function."""
    args = parse_arguments()

    # Load test cases from JSON
    print(f"Loading test cases from {args.testcases}...")
    testcases = load_testcases(args.testcases)
    print(f"Loaded {len(testcases)} test cases")

    # Create Excel file
    print(f"Creating Excel file...")

    # Create summary sheet
    summary_df = create_summary_sheet(testcases)

    # Create test cases sheet
    testcases_df = create_testcases_sheet(testcases)

    # Write to Excel
    with pd.ExcelWriter(args.output, engine='openpyxl') as writer:
        summary_df.to_excel(writer, sheet_name='サマリー', index=False, header=False)
        testcases_df.to_excel(writer, sheet_name='全テストケース', index=False)

    print(f"Excel file created at {args.output}")

    # Apply formatting
    print("Applying formatting...")
    apply_formatting(args.output)

    print(f"✅ UAT test case file generated successfully: {args.output}")
    print(f"   Total test cases: {len(testcases)}")
    print(f"   High priority: {sum(1 for tc in testcases if tc['priority'] == 'High')}")
    print(f"   Medium priority: {sum(1 for tc in testcases if tc['priority'] == 'Medium')}")
    print(f"   Low priority: {sum(1 for tc in testcases if tc['priority'] == 'Low')}")


if __name__ == '__main__':
    main()
